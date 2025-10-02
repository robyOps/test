from io import BytesIO

from django.contrib.auth.models import Group, User
from django.test import TestCase
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile

from catalog.models import Category, Priority
from tickets.models import AuditLog, Ticket, TicketAssignment
from accounts.roles import ROLE_TECH
from tickets.validators import validate_upload, UploadValidationError


class ReportsExportExcelTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="u1", password="pass")
        cat = Category.objects.create(name="Cat")
        pri = Priority.objects.create(name="Baja")
        Ticket.objects.create(
            title="Test",
            description="d",
            requester=self.user,
            category=cat,
            priority=pri,
        )

    def test_export_excel(self):
        self.client.login(username="u1", password="pass")
        url = reverse("reports_export_excel")
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        from openpyxl import load_workbook

        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertIn("Código", headers)
        first_row = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
        self.assertEqual(first_row[0], "1")


class UploadValidatorTests(TestCase):
    def test_invalid_path_rejected(self):
        class Dummy:
            name = "../../evil.txt"
            size = 1
            content_type = "text/plain"

        with self.assertRaises(UploadValidationError):
            validate_upload(Dummy())

    def test_valid_file(self):
        f = SimpleUploadedFile("ok.txt", b"x", content_type="text/plain")
        try:
            validate_upload(f)
        except UploadValidationError:
            self.fail("validate_upload raised unexpectedly")


class TicketCreationAssignmentTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(
            username="admin", email="admin@example.com", password="pass"
        )
        self.tech_group, _ = Group.objects.get_or_create(name=ROLE_TECH)
        self.tech = User.objects.create_user(
            username="tech", email="tech@example.com", password="pass"
        )
        self.tech.groups.add(self.tech_group)

        self.category = Category.objects.create(name="Cat")
        self.priority = Priority.objects.create(name="Alta")

    def test_ticket_creation_with_assignee_creates_history(self):
        self.client.login(username="admin", password="pass")
        payload = {
            "title": "Nuevo ticket",
            "description": "Descripción",
            "category": str(self.category.id),
            "priority": str(self.priority.id),
            "kind": Ticket.REQUEST,
            "assignee": str(self.tech.id),
        }

        response = self.client.post(reverse("ticket_create"), payload)
        self.assertEqual(response.status_code, 302)

        ticket = Ticket.objects.get(title="Nuevo ticket")
        self.assertEqual(ticket.assigned_to, self.tech)

        assignment = TicketAssignment.objects.get(ticket=ticket)
        self.assertEqual(assignment.from_user, self.admin)
        self.assertEqual(assignment.to_user, self.tech)

        audit_entry = AuditLog.objects.get(ticket=ticket, action="ASSIGN")
        self.assertEqual(audit_entry.actor, self.admin)
        self.assertEqual(audit_entry.meta.get("from"), None)
        self.assertEqual(audit_entry.meta.get("to"), self.tech.id)

